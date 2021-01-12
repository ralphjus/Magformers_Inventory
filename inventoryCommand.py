from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter.filedialog import askopenfilename
import tkinter.messagebox
import openpyxl
from openpyxl import load_workbook
import pickle
import Square_pull
import Square_push

class SlowCH_Manager(Canvas):
    """ 
    Manages a variable number of slow channel massages
    """
    def __init__(self,master=None,**kwargs):
        Canvas.__init__(self,master,**kwargs,width = 1000,height = 550)
        self.frame = Frame(self)
        self.create_window(0,0,anchor=N+W,window=self.frame)
        self.row = 0
        self.widgets = []
        self.max = 100
        self._init_entries()

    def _init_entries(self):
        """
        initialize the input area with labels and perhaps default values
        """
        
    def _ypos(self):
        return sum(x.winfo_reqheight() for x in self.widgets)
    
#raw_scan = str(input('''
#Welcome to super inventory manager 2k20!
#Scan item (then press enter) to add to inventory
#press b to manually enter item number from the box
#type push to push inventory to Sqaure
#type pull to pull inventory from Square
#press q to quit\n'''))
  




if __name__ == "__main__":
    workbook = ""
    catalog = ""
    root = Tk()
    root.title("Magformers Inventory Manager 2k20")
    w = 1300 # width for the Tk root
    h = 700 # height for the Tk root

    #get screen width and height
    ws = root.winfo_screenwidth() # width of the screen
    hs = root.winfo_screenheight() # height of the screen

    # calculate x and y coordinates for the Tk root window
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)

    # set the dimensions of the screen 
    # and where it is placed
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))

    try:
    # Getting back the objects:
        with open('workbook.pkl', 'rb') as f:  # Python 3: open(..., 'rb')
            workbook = pickle.load(f)
            print(workbook)
    except:
        print()
    try:
        with open('catalog.pkl', 'rb') as f:  # Python 3: open(..., 'rb')
            catalog = pickle.load(f)
            print(catalog)
    except:
        print()
#Buttons
    def Workbook():
        try:
            global workbook
            
            fname = askopenfilename(filetypes=(("Excel files", "*.xlsx"),
                                                    ("All files", "*.*") ))
            workbook = str(fname)
            tkinter.messagebox.showinfo('Success','Workbook loaded! Restart to confirm')
            # Saving the objects:
            with open('workbook.pkl', 'wb') as f:  # Python 3: open(..., 'wb')
                pickle.dump(workbook, f)

        except:
            tkinter.messagebox.showinfo('Error','Invalid sheet. Must be in .xlsx format')

    def Catalog():
        try:
            global catalog
            
            fname = askopenfilename(filetypes=(("Excel files", "*.xlsx"),
                                                    ("All files", "*.*") ))
            catalog = str(fname)
            tkinter.messagebox.showinfo('Success','Catalog loaded! Restart to confirm')
            # Saving the objects:
            with open('catalog.pkl', 'wb') as f:  # Python 3: open(..., 'wb')
                pickle.dump(catalog, f)

        except:
            tkinter.messagebox.showinfo('Error','Invalid sheet. Must be in .xlsx format')

    #boxes
    def Box(number, tab, workbook):
        wb = openpyxl.load_workbook(workbook)
        ws = wb.active
        wb.sheetnames
        sheet = wb[tab]
        hit = 0
        sku = number
        for row in range(2, sheet.max_row +1):
            if sku in str(sheet['A' + str(row)].value):
                print(sheet['B' + str(row)].value)
                sheet['E'+str(row)] = int(sheet['E'+str(row)].value) + int(sheet['C'+str(row)].value)
                wb.save(workbook)
                wb.close(workbook)
                print("Inventory count: {}".format(sheet['E'+str(row)].value))
                hit = 1
        if hit == 0:
            print("***PRODUCT NOT IN LIST***")
            add = input("Add product to inventory? y/n:\n")
            if add == "y":
                description = input("Please input description:\n")
                new_row = sheet.max_row + 1
                sheet['A'+str(new_row)] = sku
                sheet['B'+str(new_row)] = description
                sheet['E'+str(new_row)] = 1
                wb.save(workbook)
                wb.close(workbook)
                
    #Call Square syncing functions    
    def Push():
        Square_push.push_to_square()
        raw_scan = str(input())
        
    def Pull():
        Square_pull.pull_from_square()
        raw_scan = str(input())        
    #Scanner
    def Scan(description,tab,workbook):
        print(workbook)
        print(description)
        print(tab)
        wb = openpyxl.load_workbook(workbook)
        ws = wb.active
        wb.sheetnames
        sku = description
        sheet = wb[tab]
        print(description)
        print(tab)
        hit = 0
        for row in range(2, sheet.max_row +1):
            if sku in str(sheet['G' + str(row)].value):
                print(sheet['B' + str(row)].value)
                sheet['E'+str(row)] = int(sheet['E'+str(row)].value) + 1
                wb.save(workbook)
                print("Inventory count: {}".format(sheet['E'+str(row)].value))
                display = sheet['B' + str(row)].value,"Inventory count: {}".format(sheet['E'+str(row)].value)
                var.set(display)
                root.update_idletasks()
                hit = 1
        if hit == 0:
            print("***PRODUCT NOT IN LIST***")
            add = input("Add product to inventory? y/n:\n")
            if add == "y":
                description = input("Please input description:\n")
                new_row = sheet.max_row + 1
                sheet['A'+str(new_row)] = sku
                sheet['B'+str(new_row)] = description
                sheet['E'+str(new_row)] = 1
                wb.save(workbook)
            
            elif add == "n":
                print("Canceled. scan next item.")
            else:
                print("Command not recognized. please scan again and type 'y' for yes and 'n' no")
                
    all_entries = []
    #scan input
    description = tk.StringVar()
    descriptionEntry = ttk.Entry(width=13, textvariable=description)
    descriptionEntry.grid(column=0, row=0)
    #tab selection
    variable = tk.StringVar(root)
    tabs = ["Magformers","TileBlox","Clicformers","Stick-O", "Dolce"]
    variable.set(tabs[0])
    opt = tk.OptionMenu(root, variable, *tabs)
    opt.grid(row=2,column=2)
    #buttons
    b = Button(root, text = "Scan", command = lambda : Scan(description.get(),variable.get(),workbook))
    b.grid(row=2,column=0)
    c = Button(root, text = "Box", command = lambda : Box(description.get(),variable.get(),workbook))
    c.grid(row=3,column=0)
    d = Button(root, text = "Push to Square", command = lambda : Square_push.push_to_square(variable.get(),workbook,catalog))
    d.grid(row=4,column=0)
    e = Button(root, text = "Pull from Square", command = lambda : Square_pull.pull_from_square(variable.get(),workbook))
    e.grid(row =5, column = 0)
    f = Button(root, text = "Define Workbook", command = lambda : Workbook())
    f.grid(row=2,column=1)
    g = Button(root, text = "Define Catalog", command = lambda : Catalog())
    g.grid(row=3,column=1)
    if workbook != "":
        label_workbook = Label(root, text=workbook).grid(row = 6, column = 0)
    else:
        label_workbook = Label(root, text="No workbook loaded!").grid(row = 6, column = 0)
    if catalog != "":
        label_catalog = Label(root, text=catalog).grid(row = 7, column = 0)
    else:
        label_workbook = Label(root, text="No catalog loaded!").grid(row = 7, column = 0)
    var = StringVar()
    var.set('Items will appear here once entered')

    h = Label(root,fg="green", textvariable = var)
    h.grid(row=8,column=0)

    root.mainloop()
