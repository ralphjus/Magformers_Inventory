import openpyxl
from openpyxl import load_workbook
from square.client import Client
import datetime 
from datetime import timezone
import uuid
import tkinter as tk
from tkinter import ttk

def push_to_square(tab,workbook,catalog):
    wb = openpyxl.load_workbook(catalog)
    ws = wb.active
    wb.sheetnames
    sheet = wb['Items']

    item_ids = []

    popup = tk.Toplevel()
    tk.Label(popup, text="Pushing to Square...").grid(row=0,column=0)
    progress = 0
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(popup, variable=progress_var, maximum=100)
    progress_bar.grid(row=1,column=0)

    for item in range(2,sheet.max_row +1):
        item_ids.append(sheet['A'+str(item)].value)

    client = Client(
        access_token='ACCESSTOKEN',
        environment='production',
    )

    result = client.catalog.list_catalog(
      types = "ITEM_VARIATION"
    )

    if result.is_success():
        print()
        #print(result.body)
    elif result.is_error():
        print(result.errors)

    inventory_api = client.inventory

    body = {}
    body['catalog_object_ids'] = item_ids
    body['location_ids'] = ['CMRZVV9YNFAJP']

    result = inventory_api.batch_retrieve_inventory_counts(body)

    if result.is_success():
        req = result.body
        item_list = req['counts']

    else:
        print(result.errors)

    d = datetime.datetime.utcnow()

    wb = openpyxl.load_workbook(workbook)
    ws = wb.active
    wb.sheetnames
    sheet = wb[tab]
    
    progress_step = float(100.0/sheet.max_row)
    for item in range(4,sheet.max_row +1):
        popup.update()
        if sheet['F'+str(item)].value != None:

            body = {}
            body['idempotency_key'] =  str(uuid.uuid4())
            body['changes'] = []
            body['changes'].append({})
            body['changes'][0]['type'] = 'PHYSICAL_COUNT'
            body['changes'][0]['physical_count'] = {}
            body['changes'][0]['physical_count']['catalog_object_id'] = str(sheet['F'+str(item)].value)
            body['changes'][0]['physical_count']['state'] = 'IN_STOCK'
            body['changes'][0]['physical_count']['location_id'] = 'CMRZVV9YNFAJP'
            body['changes'][0]['physical_count']['quantity'] = str(sheet['E'+str(item)].value)
            body['changes'][0]['physical_count']['occurred_at'] = d.isoformat("T") + "Z"
            body['ignore_unchanged_counts'] = True

            result = inventory_api.batch_change_inventory(body)

            if result.is_success():
                print(result.body)
            elif result.is_error():
                print(result.errors)
        progress += progress_step
        progress_var.set(progress)
    tk.Label(popup, text="Success!").grid(row=3,column=0)
    popup.update()
    
if __name__ == '__main__':
    print("You should run me from Inventory command!")
