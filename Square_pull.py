import openpyxl
from openpyxl import load_workbook
from square.client import Client
import datetime 
from datetime import timezone
import uuid
import json
import tkinter as tk
from tkinter import ttk

def pull_from_square(tab,workbook):

    wb = openpyxl.load_workbook(workbook)
    ws = wb.active
    wb.sheetnames
    sheet = wb[tab]

    item_ids = []
    
    popup = tk.Toplevel()
    tk.Label(popup, text="Pulling from Square...").grid(row=0,column=0)

    progress = 0
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(popup, variable=progress_var, maximum=100)
    progress_bar.grid(row=1,column=0)
    
    for item in range(4,sheet.max_row +1):
        item_ids.append(sheet['F'+str(item)].value)

    client = Client(
        access_token='ACCESSTOKEN',
        environment='production',
    )
    objects = []
    quantities = []
    
    progress_step = float(100.0/len(item_ids))
    for item in item_ids:
        popup.update()
        if item != None:
            result = client.inventory.retrieve_inventory_count(
            catalog_object_id = item,
            location_ids = "CMRZVV9YNFAJP"
        )

            if result.is_success():
                json_object = json.dumps(result.body, indent = 4)
                try:
                    data = json.loads(json_object)
                    counts = data['counts']
                    object_id = counts[0]['catalog_object_id']
                    quantity = counts[0]['quantity']
                    objects.append(object_id)
                    quantities.append(quantity)
                except:
                    print("oops!!!")
                    objects.append(0)
                    quantities.append(0)
                
            elif result.is_error():
              print(result.errors)
          
        else:
            objects.append(0)
            quantities.append(0)
        progress += progress_step
        progress_var.set(progress)

    tk.Label(popup, text="Updating workbook...").grid(row=0,column=0)
    progress = 0
    progress_var.set(progress)
    popup.update()
    progress_step = float(100.0/len(objects))
    for item in range(len(objects)):
        popup.update()
        if objects[item] in item_ids:
            index = int(item_ids.index(objects[item]))
            row = index + 4
            sheet['E'+str(row)] = int(quantities[item])
            print("{} is now {}".format(sheet['B'+str(row)].value,quantities[item]))
        progress += progress_step
        progress_var.set(progress)
    wb.save(workbook)
    tk.Label(popup, text="Success!").grid(row=2,column=0)
    popup.update()
    
if __name__ == '__main__':
    print("You should run me from Inventory command!")
